import openpyxl
wb = openpyxl.load_workbook('data/infracoop_bd.xlsx', data_only=True)
ws = wb['Datasets']
print('Total filas:', ws.max_row)
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
    print(f'Fila {i}:', [str(v)[:30] if v else None for v in row[:5]])
