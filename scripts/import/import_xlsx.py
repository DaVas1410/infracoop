"""
Imports infracoop_bd.xlsx into Supabase.
Usage: python import_xlsx.py [--validate-urls] [--dry-run]

Reads sheets: Datasets (42 rows), Normativas (35 rows).
Skips sheet: Metodología (documentation only).
"""
import argparse
import json
import os
import sys
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv
from supabase import create_client

# Load env from repo root .env
load_dotenv(Path(__file__).parent.parent.parent / '.env')

SUPABASE_URL = os.environ['VITE_SUPABASE_URL']
SUPABASE_KEY = os.environ['VITE_SUPABASE_ANON_KEY']
XLSX_PATH = Path(__file__).parent.parent.parent / 'data' / 'infracoop_bd.xlsx'

KNOWN_PROBLEMATIC = {
    'DS-006': 'PDF — no machine-readable',
    'DS-012': 'Power BI — not downloadable',
    'DS-021': 'DBF format',
    'DS-022': 'DBF format',
    'DS-023': 'DBF format',
    'DS-029': 'PDF — no machine-readable',
    'DS-030': 'Power BI — not downloadable',
    'DS-031': 'Power BI — not downloadable',
    'DS-035': 'Discontinued July 2023',
}


def parse_agendas(cell_value):
    if not cell_value:
        return []
    return [a.strip() for a in str(cell_value).split('·') if a.strip()]


def validate_url(url, dataset_id):
    if not url:
        return False
    if dataset_id in KNOWN_PROBLEMATIC:
        return False
    try:
        r = requests.head(url, timeout=8, allow_redirects=True)
        return r.status_code < 400
    except Exception:
        return False


def import_datasets(ws, client, validate_urls, dry_run):
    log = []
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    for row in rows:
        if not row[0]:
            continue
        (ds_id, titulo, fuente, pais, anio, subtema, agendas_raw,
         frecuencia, desagregacion, accesibilidad, url, descripcion) = row[:12]

        url_valida = validate_url(url, str(ds_id)) if validate_urls else True

        record = {
            'id': str(ds_id),
            'titulo': str(titulo) if titulo else '',
            'fuente_organismo': str(fuente) if fuente else None,
            'pais_iso3': str(pais) if pais else None,
            'anio_publicacion': int(anio) if anio else None,
            'subtema': str(subtema) if subtema else None,
            'agendas': parse_agendas(str(agendas_raw) if agendas_raw else ''),
            'frecuencia': str(frecuencia) if frecuencia else None,
            'desagregacion_geo': str(desagregacion) if desagregacion else None,
            'accesibilidad_formato': str(accesibilidad) if accesibilidad else None,
            'url_descarga': str(url) if url else None,
            'url_valida': url_valida,
            'descripcion_notas': str(descripcion) if descripcion else None,
            'es_sintetico': False,
        }

        status = 'skipped (dry-run)' if dry_run else 'ok'
        if not dry_run:
            try:
                client.table('datasets').upsert(record).execute()
            except Exception as e:
                status = f'error: {e}'

        log.append({'id': str(ds_id), 'titulo': str(titulo), 'status': status,
                    'url_valida': url_valida,
                    'nota': KNOWN_PROBLEMATIC.get(str(ds_id), '')})
        titulo_str = str(titulo)[:50] if titulo else '(sin título)'
        print(f"  {'✓' if 'error' not in status else '✗'} {ds_id}: {titulo_str}")

    return log


def import_normativas(ws, client, dry_run):
    log = []
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    for row in rows:
        if not row[0]:
            continue
        (nm_id, nombre, organismo, tipo, pais_alcance, anio,
         articulo, obligacion, agendas_raw, url, descripcion) = row[:11]

        record = {
            'id': str(nm_id),
            'nombre': str(nombre) if nombre else '',
            'organismo_emisor': str(organismo) if organismo else None,
            'tipo': str(tipo) if tipo else None,
            'pais_alcance': str(pais_alcance) if pais_alcance else None,
            'anio_adopcion': int(anio) if anio else None,
            'articulo_numeral': str(articulo) if articulo else None,
            'obligacion_datos': str(obligacion) if obligacion else None,
            'agendas': parse_agendas(str(agendas_raw) if agendas_raw else ''),
            'url_texto_oficial': str(url) if url else None,
            'descripcion_notas': str(descripcion) if descripcion else None,
            'es_sintetico': False,
        }

        status = 'skipped (dry-run)' if dry_run else 'ok'
        if not dry_run:
            try:
                client.table('normativas').upsert(record).execute()
            except Exception as e:
                status = f'error: {e}'

        log.append({'id': str(nm_id), 'nombre': str(nombre), 'status': status})
        print(f"  {'✓' if 'error' not in status else '✗'} {nm_id}: {nombre[:50]}")

    return log


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--validate-urls', action='store_true',
                        help='Send HTTP HEAD request to validate each dataset URL')
    parser.add_argument('--dry-run', action='store_true',
                        help='Parse Excel without writing to Supabase')
    args = parser.parse_args()

    if not XLSX_PATH.exists():
        print(f'ERROR: {XLSX_PATH} not found', file=sys.stderr)
        sys.exit(1)

    client = create_client(SUPABASE_URL, SUPABASE_KEY)
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    print('\n=== Importando Datasets ===')
    datasets_log = import_datasets(wb['Datasets'], client, args.validate_urls, args.dry_run)

    print('\n=== Importando Normativas ===')
    normativas_log = import_normativas(wb['Normativas'], client, args.dry_run)

    log = {'datasets': datasets_log, 'normativas': normativas_log}
    log_path = Path(__file__).parent / 'import_log.json'
    log_path.write_text(json.dumps(log, ensure_ascii=False, indent=2))

    ok_d = sum(1 for r in datasets_log if r['status'] in ('ok', 'skipped (dry-run)'))
    ok_n = sum(1 for r in normativas_log if r['status'] in ('ok', 'skipped (dry-run)'))
    print(f'\n✓ Listo. Datasets: {ok_d}/{len(datasets_log)}  Normativas: {ok_n}/{len(normativas_log)}')
    print(f'  Log guardado en {log_path}')


if __name__ == '__main__':
    main()
